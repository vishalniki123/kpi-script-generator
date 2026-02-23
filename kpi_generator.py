import os
import re
import shutil
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font

FONT = Font(name="Arial", size=10)

# ================= COLUMN MAPS =================
COLUMN_MAP_PSQ = {
    "KPINAME": 1,
    "PARAM": 4,
    "FORMULA": 6,
    "FORMAT": 7,
    "TARGET": 9,
    "LOGIC": 10
}

COLUMN_MAP_SQD = {
    "KPINAME": 1,
    "PARAM": 4,
    "FORMULA": 7,
    "FORMAT": 8,
    "TARGET": 11,
    "LOGIC": 12,
    "QUARTERLY": 10
}

# ================= SAFE FILENAME =================
def normalize_excel_operators(text):
    replacements = {
        "<=": "≤",
        ">=": "≥",
        "<": "＜",
        ">": "＞"
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text


def make_safe_filename(name):
    name = normalize_excel_operators(name)
    name = unicodedata.normalize("NFC", name)
    name = re.sub(r'[\\/:*?"|]', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def get_unique_filepath(directory, filename):
    base, ext = os.path.splitext(filename)
    counter = 1
    path = os.path.join(directory, filename)

    while os.path.exists(path):
        path = os.path.join(directory, f"{base}_{counter}{ext}")
        counter += 1

    return path


# ================= FILE CREATION =================
def create_kpi_file(
        output_dir, template, kpi_name, formula, fmt,
        target, logic, params, quarterly_logic
):
    safe_name = make_safe_filename(kpi_name)
    out_file = get_unique_filepath(output_dir, f"{safe_name}.xlsx")

    shutil.copy(template, out_file)

    wb = load_workbook(out_file)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = FONT
                if isinstance(cell.value, str):
                    cell.value = (
                        cell.value
                        .replace("KPINAME", kpi_name)
                        .replace("KPIFORMULA", formula)
                        .replace("KPIFORMAT", fmt)
                        .replace("KPITARGET", target)
                        .replace("KPILOGIC", logic)
                        .replace("KPIPARAMETER", params)
                        .replace("QUARTERLYLOGIC", quarterly_logic)
                    )

    wb.save(out_file)


# ================= SHEET PROCESSOR =================
def process_sheet(
        wb, sheet_index, output_dir,
        normal_template, tracking_template,
        column_map
):
    ws = wb.worksheets[sheet_index]

    current_kpi = None
    formula = fmt = target = logic = quarterly = ""
    params = []

    for row in ws.iter_rows(min_row=2):
        kpi_name = row[column_map["KPINAME"]].value
        param = row[column_map["PARAM"]].value
        f = row[column_map["FORMULA"]].value
        frmt = row[column_map["FORMAT"]].value
        tgt = row[column_map["TARGET"]].value
        rpt = row[column_map["LOGIC"]].value
        ql = row[column_map["QUARTERLY"]].value if "QUARTERLY" in column_map else ""

        if kpi_name:
            if current_kpi:
                template = tracking_template if target.lower() == "tracking only" else normal_template
                create_kpi_file(
                    output_dir, template,
                    current_kpi, formula, fmt,
                    target, logic,
                    " & ".join(params),
                    quarterly
                )

            current_kpi = str(kpi_name).strip()
            formula = str(f).strip() if f else ""
            fmt = str(frmt).strip() if frmt else ""
            target = str(tgt).strip() if tgt else ""
            logic = str(rpt).strip() if rpt else ""
            quarterly = str(ql).strip() if ql else ""
            params = []

        if param:
            params.append(str(param).strip())

    if current_kpi:
        template = tracking_template if target.lower() == "tracking only" else normal_template
        create_kpi_file(
            output_dir, template,
            current_kpi, formula, fmt,
            target, logic,
            " & ".join(params),
            quarterly
        )


# ================= ENTRY POINT FOR UI =================
def run_generator(
        kpi_excel,
        psq_script_1,
        psq_script_2,
        sqd_script_1,
        sqd_script_2,
        psq_output,
        sqd_output
):
    os.makedirs(psq_output, exist_ok=True)
    os.makedirs(sqd_output, exist_ok=True)

    wb = load_workbook(kpi_excel, data_only=True)

    process_sheet(
        wb,
        sheet_index=0,
        output_dir=psq_output,
        normal_template=psq_script_1,
        tracking_template=psq_script_2,
        column_map=COLUMN_MAP_PSQ
    )

    process_sheet(
        wb,
        sheet_index=1,
        output_dir=sqd_output,
        normal_template=sqd_script_1,
        tracking_template=sqd_script_2,
        column_map=COLUMN_MAP_SQD
    )